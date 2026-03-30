"""
Automated Excel report generator.

Produces a multi-sheet Excel report with dashboard, historical data,
regime analysis, stress testing results, and correlation matrices.
Designed as the Monday morning briefing for a credit analyst.
"""

import logging
from pathlib import Path
from datetime import datetime

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

from src.config import CONFIG, PROJECT_ROOT

logger = logging.getLogger(__name__)

# --- Style constants ---
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="2C3E50")
SUBHEADER_FILL = PatternFill("solid", fgColor="34495E")
RED_FILL = PatternFill("solid", fgColor="FADBD8")
RED_FONT = Font(color="C0392B")
GREEN_FILL = PatternFill("solid", fgColor="D5F5E3")
GREEN_FONT = Font(color="27AE60")
NEUTRAL_FILL = PatternFill("solid", fgColor="F8F9FA")
THIN_BORDER = Border(
    left=Side(style="thin", color="D5D8DC"),
    right=Side(style="thin", color="D5D8DC"),
    top=Side(style="thin", color="D5D8DC"),
    bottom=Side(style="thin", color="D5D8DC"),
)


def write_header_row(ws, row: int, headers: list[str], start_col: int = 1) -> None:
    """Write a formatted header row."""
    for i, header in enumerate(headers):
        cell = ws.cell(row=row, column=start_col + i, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def auto_column_width(ws, min_width: int = 10, max_width: int = 30) -> None:
    """Auto-adjust column widths based on content."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        adjusted = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def apply_zscore_formatting(ws, col_letter: str, start_row: int, end_row: int) -> None:
    """Apply conditional formatting to z-score cells."""
    for row in range(start_row, end_row + 1):
        cell = ws[f"{col_letter}{row}"]
        if cell.value is None:
            continue
        try:
            val = float(cell.value)
        except (ValueError, TypeError):
            continue

        if val > 2:
            cell.fill = RED_FILL
            cell.font = RED_FONT
        elif val > 1:
            cell.fill = PatternFill("solid", fgColor="FDEBD0")
        elif val < -1:
            cell.fill = GREEN_FILL
            cell.font = GREEN_FONT


def create_dashboard_sheet(ws, df: pd.DataFrame) -> None:
    """
    Create the Dashboard sheet — the analyst's Monday morning view.

    Shows current spread levels, changes, z-scores, percentiles,
    and the current market regime.
    """
    # Title
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = f"Credit Spread Dashboard — {df.index[-1].date()}"
    title_cell.font = Font(bold=True, size=14, color="2C3E50")
    title_cell.alignment = Alignment(horizontal="center")

    # Current regime
    ws["A3"] = "Current Regime:"
    ws["A3"].font = Font(bold=True, size=12)
    ws["B3"] = df["regime"].iloc[-1].upper()
    ws["B3"].font = Font(bold=True, size=12)
    regime = df["regime"].iloc[-1]
    if regime == "crisis":
        ws["B3"].font = Font(bold=True, size=12, color="C0392B")
    elif regime == "risk_on":
        ws["B3"].font = Font(bold=True, size=12, color="27AE60")
    else:
        ws["B3"].font = Font(bold=True, size=12, color="F39C12")

    ws["C3"] = f"(probability: {df['regime_proba'].iloc[-1]:.1%})"
    ws["C3"].font = Font(italic=True, color="7F8C8D")

    # Spread metrics table
    headers = [
        "Rating", "Current (%)", "Δ1d", "Δ5d", "Δ20d",
        "Z-Score 60d (EMA)", "Z-Score 252d (EMA)",
        "Percentile 252d", "Vol 20d",
    ]
    write_header_row(ws, 5, headers)

    spread_info = [
        ("AAA", "aaa_spread"),
        ("AA", "aa_spread"),
        ("BBB", "bbb_spread"),
        ("HY", "hy_spread"),
    ]

    last = df.iloc[-1]
    for i, (label, col) in enumerate(spread_info):
        row = 6 + i
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=1).border = THIN_BORDER

        # Current level
        ws.cell(row=row, column=2, value=round(last.get(col, 0), 4))

        # Changes
        for j, window in enumerate([1, 5, 20]):
            change_col = f"{col}_change_{window}d"
            val = last.get(change_col, None)
            cell = ws.cell(row=row, column=3 + j, value=round(val, 4) if val else None)
            cell.border = THIN_BORDER

        # Z-scores (EMA)
        for j, window in enumerate([60, 252]):
            zscore_col = f"{col}_zscore_ema_{window}d"
            val = last.get(zscore_col, None)
            cell = ws.cell(row=row, column=6 + j, value=round(val, 2) if val else None)
            cell.border = THIN_BORDER

        # Percentile
        pctile_col = f"{col}_pctile_252d"
        val = last.get(pctile_col, None)
        cell = ws.cell(row=row, column=8, value=round(val, 2) if val else None)
        cell.border = THIN_BORDER

        # Volatility
        vol_col = f"{col}_rolling_vol_20d"
        val = last.get(vol_col, None)
        cell = ws.cell(row=row, column=9, value=round(val, 4) if val else None)
        cell.border = THIN_BORDER

    # Apply z-score formatting
    for col_letter in ["F", "G"]:
        apply_zscore_formatting(ws, col_letter, 6, 9)

    # Compression ratio
    ws["A12"] = "BBB/HY Ratio:"
    ws["A12"].font = Font(bold=True)
    ws["B12"] = round(last.get("bbb_hy_ratio", 0), 4)

    # Note about z-scores
    ws["A14"] = "Note: Z-scores use EMA (no ghost effect). Percentiles are the primary metric."
    ws["A14"].font = Font(italic=True, size=9, color="7F8C8D")
    ws["A15"] = "Z-score > 2 (red) = historically wide. Z-score < -1 (green) = historically tight."
    ws["A15"].font = Font(italic=True, size=9, color="7F8C8D")

    auto_column_width(ws)
    logger.info("Created Dashboard sheet")





def create_history_sheet(ws, df: pd.DataFrame) -> None:
    """
    Create the Historical sheet — last 252 trading days of spread data
    plus embedded charts.
    """
    from openpyxl.drawing.image import Image as XLImage

    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = "Historical Spread Data (Last 252 Trading Days)"
    ws["A1"].font = Font(bold=True, size=12, color="2C3E50")

    # Data table
    last_252 = df.tail(252)
    data_cols = ["aaa_spread", "aa_spread", "bbb_spread", "hy_spread", "vix"]
    available_cols = [c for c in data_cols if c in last_252.columns]

    headers = ["Date"] + [c.replace("_spread", "").upper() for c in available_cols]
    write_header_row(ws, 3, headers)

    for i, (date, row) in enumerate(last_252[available_cols].iterrows()):
        ws.cell(row=4 + i, column=1, value=date.strftime("%Y-%m-%d"))
        ws.cell(row=4 + i, column=1).border = THIN_BORDER
        for j, col in enumerate(available_cols):
            cell = ws.cell(row=4 + i, column=2 + j, value=round(row[col], 4))
            cell.border = THIN_BORDER

    # Embed charts (if figures exist)
    figures_dir = PROJECT_ROOT / "figures"
    chart_col = len(available_cols) + 3  # place charts to the right of data

    charts_to_embed = [
        "spread_history.png",
        "zscore_bbb.png",
        "volatility_comparison.png",
    ]

    row_offset = 3
    for chart_file in charts_to_embed:
        chart_path = figures_dir / chart_file
        if chart_path.exists():
            try:
                img = XLImage(str(chart_path))
                img.width = 600
                img.height = 300
                cell_ref = f"{get_column_letter(chart_col)}{row_offset}"
                ws.add_image(img, cell_ref)
                row_offset += 18  # space between charts
            except Exception as e:
                logger.warning("Could not embed %s: %s", chart_file, e)

    auto_column_width(ws)
    logger.info("Created Historical sheet")



def create_regime_sheet(
    ws,
    transition_matrix: pd.DataFrame,
    regime_stats: pd.DataFrame,
) -> None:
    """
    Create the Regime Analysis sheet — transition matrix and statistics.
    """
    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = "Regime Analysis — Hidden Markov Model"
    ws["A1"].font = Font(bold=True, size=12, color="2C3E50")

    # Transition matrix
    ws["A3"] = "Transition Probability Matrix"
    ws["A3"].font = Font(bold=True, size=11)

    # Headers
    regimes = list(transition_matrix.columns)
    write_header_row(ws, 4, ["From \\ To"] + regimes)

    for i, (from_regime, row) in enumerate(transition_matrix.iterrows()):
        ws.cell(row=5 + i, column=1, value=from_regime)
        ws.cell(row=5 + i, column=1).font = Font(bold=True)
        ws.cell(row=5 + i, column=1).border = THIN_BORDER
        for j, to_regime in enumerate(regimes):
            val = row[to_regime]
            cell = ws.cell(row=5 + i, column=2 + j, value=round(val, 4))
            cell.number_format = "0.00%"
            cell.border = THIN_BORDER
            # Color: darker for higher probability
            if val > 0.8:
                cell.fill = PatternFill("solid", fgColor="E74C3C")
                cell.font = Font(color="FFFFFF")
            elif val > 0.5:
                cell.fill = PatternFill("solid", fgColor="F39C12")

    # Regime stats
    start_row = 5 + len(transition_matrix) + 2
    ws.cell(row=start_row, column=1, value="Regime Statistics")
    ws.cell(row=start_row, column=1).font = Font(bold=True, size=11)

    stat_headers = list(regime_stats.columns)
    write_header_row(ws, start_row + 1, stat_headers)

    for i, (_, row) in enumerate(regime_stats.iterrows()):
        for j, col in enumerate(stat_headers):
            val = row[col]
            if isinstance(val, float):
                val = round(val, 4)
            cell = ws.cell(row=start_row + 2 + i, column=1 + j, value=val)
            cell.border = THIN_BORDER

    auto_column_width(ws)
    logger.info("Created Regime Analysis sheet")




def create_stress_sheet(
    ws,
    stress_results: pd.DataFrame,
    stress_confidence: pd.DataFrame | None = None,
) -> None:
    """
    Create the Stress Testing sheet — results by scenario and horizon.
    Includes max widening (the real risk metric) and confidence intervals.
    """
    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = "Historical Stress Testing Results"
    ws["A1"].font = Font(bold=True, size=12, color="2C3E50")

    ws["A2"] = "Max Widening is the primary risk metric — it captures intra-window extremes that trigger margin calls."
    ws["A2"].font = Font(italic=True, size=9, color="7F8C8D")

    # Results table
    headers = list(stress_results.columns)
    write_header_row(ws, 4, headers)

    for i, (_, row) in enumerate(stress_results.iterrows()):
        for j, col in enumerate(headers):
            val = row[col]
            if isinstance(val, float):
                val = round(val, 4)
            cell = ws.cell(row=5 + i, column=1 + j, value=val)
            cell.border = THIN_BORDER

            # Flag low sample sizes
            if col == "n_episodes" and isinstance(val, (int, float)) and val < 10:
                cell.fill = PatternFill("solid", fgColor="FDEBD0")
                cell.font = Font(color="E67E22", italic=True)

    # Confidence intervals (if available)
    if stress_confidence is not None and not stress_confidence.empty:
        ci_start = 5 + len(stress_results) + 2
        ws.cell(row=ci_start, column=1, value="Bootstrap 95% Confidence Intervals (Max Widening)")
        ws.cell(row=ci_start, column=1).font = Font(bold=True, size=11)

        ci_headers = list(stress_confidence.columns)
        write_header_row(ws, ci_start + 1, ci_headers)

        for i, (_, row) in enumerate(stress_confidence.iterrows()):
            for j, col in enumerate(ci_headers):
                val = row[col]
                if isinstance(val, float):
                    val = round(val, 4)
                cell = ws.cell(row=ci_start + 2 + i, column=1 + j, value=val)
                cell.border = THIN_BORDER

                # Highlight non-significant results
                if col == "significant" and val is False:
                    cell.fill = RED_FILL
                    cell.font = RED_FONT

    auto_column_width(ws)
    logger.info("Created Stress Testing sheet")


    